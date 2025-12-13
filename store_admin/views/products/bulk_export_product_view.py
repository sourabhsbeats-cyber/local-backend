import pandas as pd
from openpyxl import load_workbook
from openpyxl import load_workbook
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl import load_workbook  # pip install openpyxl
from openpyxl import Workbook

headers = ['Brand', 'SKU', 'Title', 'Subtitle', 'Description', 'Short Desc',
                    'Manufacturer', 'EAN', 'UPC', 'ISBN', 'MPN', 'Country of Origin', 'Condition',
                    'ASIN', 'FNSKU', 'FBA SKU', 'FBA', 'Amazon Size', 'Barcode Label Type', 'Prep Type',
                    'Stock Status', 'Status', 'Publish', 'Warranty', 'Product Tags', 'Sales Price',
                    'Retail Price (RRP)', 'Cost per Item', 'Product Margin %', 'Profit', 'Minimum Price',
                    'Maximum Price', 'Estimated Shipping Cost', 'Preferred Vendor', 'Vendor SKU', 'Parent SKU',
                    'Colour', 'Size', 'Material', 'Compatibility', 'Height (cm)', 'Width (cm)', 'Depth (cm)',
                    'Weight (kg)', 'Fast Dispatch', 'Free Shipping', 'Is the product bulky?', 'Is Taxable', 'Shipping Logic',
                    'International Note', 'Example Reference (optional)', 'Ships from', 'Handling Time (days)',
                    'SBAU', 'Local 3PL', 'China', 'USA', 'Main Image']

cond_options = [
        {"label": "new",          "value": "New"},
        {"label": "used",         "value": "Used"},
        {"label": "refurbished",  "value": "Refurbished"},
        {"label": "open_box",     "value": "Open Box"},
    ]

barcode_options = [
    {"label": "manufacturer",   "value": "Manufacturer"},
    {"label": "amazon_barcode", "value": "Amazon Barcode"},
]
amazon_size_optns = [
    {"label": "standard",   "value": "Standard"},
    {"label": "oversize", "value": "Oversize"},
]
prep_type_options = [
    {"label": "no_prep_needed", "value": "No Prep Needed"},
    {"label": "polybag",        "value": "Polybagging"},
    {"label": "bubble_wrap",    "value": "Bubble Wrap"},
    {"label": "labeling",       "value": "Labeling"},
]

stock_sts_options = [
    {"label": 1, "value": "In Stock"},
    {"label": 2, "value": "Out of Stock"},
    {"label": 3, "value": "Preorder"},
    {"label": 4, "value": "Discontinued"},
]

status_options = [
    {"label": 1, "value": "Active"},
    {"label": 2, "value": "Draft"},
    {"label": 3, "value": "Discontinued"},
    {"label": 4, "value": "Archived"},
]

pub_status_options = [
    {"label": 1, "value": "Yes"},
    {"label": 0, "value": "No"},
]
free_ship_options = [
    {"label": 1, "value": "Yes"},
    {"label": 0, "value": "No"},
]
export_prod_sql = """
                   SELECT b.name AS "Brand",
	p.sku AS "SKU",
	p.title AS "Title",
	p.subtitle AS "Subtitle",
	p.description AS "Description",
	p.short_description AS "Short Desc",
	m.name AS "Manufacturer",
	p.ean AS "EAN",
	p.upc AS "UPC",
	p.isbn AS "ISBN",
	p.mpn AS "MPN",
	c.name  AS "Country of Origin",
	p.status_condition AS "Condition",
	p.asin AS "ASIN",
	p.fnsku AS "FNSKU",
	p.fba_sku AS "FBA SKU",
	p.is_fba AS "FBA",
	p.amazon_size AS "Amazon Size",
	p.barcode_label_type AS "Barcode Label Type",
	p.prep_type AS "Prep Type",
	p.stock_status AS "Stock Status",
	p.status AS "Status",
	p.publish_status AS "Publish",
	p.warranty AS "Warranty",
	p.product_tags AS "Product Tags",
	pp.sale_price AS "Sales Price",
	pp.retail_price AS "Retail Price (RRP)",
	pp.cost_per_item AS "Cost per Item",
	pp.margin_percent AS "Product Margin %",
	pp.profit AS "Profit",
	pp.min_price AS "Minimum Price",
	pp.max_price AS "Maximum Price",
	pp.estimated_shipping_cost AS "Estimated Shipping Cost",
	v.display_name  AS "Preferred Vendor",
	p.vendor_sku AS "Vendor SKU",
	ps.attrib_color AS "Colour",
	ps.attrib_size AS "Size",
	ps.attrib_material AS "Material",
	ps.attrib_compatibility AS "Compatibility",
	ps.attrib_height AS "Height (cm)",
	ps.attrib_width AS "Width (cm)",
	ps.attrib_depth AS "Depth (cm)",
	ps.attrib_weight AS "Weight (kg)",
	sh.fast_dispatch AS "Fast Dispatch",
	sh.free_shipping AS "Free Shipping",
	sh.bulky_product AS "Is the product bulky?",
	sh.international_note AS "International Note",
	sh.example_reference AS "Example Reference (optional)",
	sh.ships_from AS "Ships from",
	sh.handling_time_days AS "Handling Time (days)",
	COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'China')
              AND product_id = p.product_id
       ), 0
    ) AS "China",

    COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'Local 3PL')
              AND product_id = p.product_id
       ), 0
    ) AS "Local 3PL",

    COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'SBAU')
              AND product_id = p.product_id
       ), 0
    ) AS "SBAU",

    COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'USA')
              AND product_id = p.product_id
       ), 0
    ) AS "USA",
    p.main_image AS "Main Image",
    p.is_taxable AS "Is Taxable"
	FROM store_admin_product p
	 LEFT JOIN store_admin_brand b ON b.brand_id = p.brand_id
	 LEFT JOIN store_admin_manufacturer m ON m.manufacturer_id = p.manufacturer_id
	 LEFT JOIN store_admin_countries c ON c.id = p.country_origin_id
	 LEFT JOIN store_admin_vendor v ON v.id = p.preferred_vendor
	 LEFT JOIN store_admin_product_price_details pp ON pp.product_id = p.product_id
	 LEFT JOIN store_admin_product_static_attributes ps ON ps.product_id = p.product_id
	 LEFT JOIN store_admin_product_shipping_details sh ON sh.product_id = p.product_id
	 LEFT JOIN store_admin_product_warehouse pw1 ON pw1.product_id = p.product_id AND pw1.warehouse_id = 1
	 LEFT JOIN store_admin_product_warehouse pw2 ON pw2.product_id = p.product_id AND pw2.warehouse_id = 2
	 LEFT JOIN store_admin_product_warehouse pw3 ON pw3.product_id = p.product_id AND pw3.warehouse_id = 3
ORDER BY p.product_id ASC
                    """
def map_option(value, options):
    if value is None or value == "":
        return ""
        # If input is integer → match directly without lower()
    if isinstance(value, int):
        return next((o["value"] for o in options if o["value"] == value), "")
        # else process as string
    v = str(value).strip().lower()
    return next((o["value"] for o in options if o["value"].lower() == v), "")


def map_int_option(val, options):
    if val is None:   # ✅ handles NULL safely
        return ""
    return next((o["value"] for o in options if o["label"] == val), "")

def map_label_option(val, options):
    # handle null/empty safely for CSV
    if val is None or val == "":
        return ""

    # int input → match label directly
    if isinstance(val, int):
        return next((o["value"] for o in options if o["label"] == val), "")

    # string input → compare with option.label case-insensitive
    v = str(val).strip().lower()
    return next(
        (o["value"] for o in options if str(o["label"]).strip().lower() == v),
        ""
    )

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import connection

def export_products_xlsx(request):
    with connection.cursor() as cur:
        cur.execute(export_prod_sql)
        cols = [c[0] for c in cur.description]  # ✅ get column names dynamically
        rows = cur.fetchall()
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="Products.csv"'

        writer = csv.writer(response)

        # Write header row exactly as received from SQL
        writer.writerow(cols)

        #  Write each data row safely
        for r in rows:
            obj = dict(zip(cols, r))  # tuple → dict mapping
            writer.writerow([
                obj.get("Brand") or "",
                obj.get("SKU") or "",
                obj.get("Title") or "",
                obj.get("Subtitle") or "",
                obj.get("Description") or "",
                obj.get("Short Desc") or "",
                obj.get("Manufacturer") or "",
                obj.get("EAN") or "",
                obj.get("UPC") or "",
                obj.get("ISBN") or "",
                obj.get("MPN") or "",
                obj.get("Country of Origin") or "",
                map_option(obj.get("Condition"), cond_options),
                obj.get("ASIN") or "",
                obj.get("FNSKU") or "",
                obj.get("FBA SKU") or "",
                "Yes" if obj.get("FBA") else "No",
                map_label_option(obj.get("Amazon Size"), amazon_size_optns),
                map_label_option(obj.get("Barcode Label Type"), barcode_options),
                map_label_option(obj.get("Prep Type"), prep_type_options),
                map_label_option(obj.get("Stock Status"), stock_sts_options),
                map_label_option(obj.get("Status"), status_options),
                map_label_option(obj.get("Publish"), pub_status_options),
                (str(obj.get("Warranty")).strip() + " Months") if obj.get("Warranty") not in [None, ""] else "",
                obj.get("Product Tags") or "",
                obj.get("Sales Price") or "",
                obj.get("Retail Price (RRP)") or "",
                obj.get("Cost per Item") or "",
                obj.get("Product Margin %") or "",
                obj.get("Profit") or "",
                obj.get("Minimum Price") or 0,
                obj.get("Maximum Price") or 0,
                obj.get("Estimated Shipping Cost") or 0,
                obj.get("Preferred Vendor") or "",
                obj.get("Vendor SKU") or "",
                #obj.get("") or "", #Parent SKU
                obj.get("Colour") or "",
                obj.get("Size") or "",
                obj.get("Material") or "",
                obj.get("Compatibility") or "",
                obj.get("Height (cm)") or "",
                obj.get("Width (cm)") or "",
                obj.get("Depth (cm)") or "",
                obj.get("Weight (kg)") or "",
                map_label_option(obj.get("Fast Dispatch"), pub_status_options),
                map_label_option(obj.get("Free Shipping"), pub_status_options),
                map_label_option(obj.get("Is the product bulky?"), pub_status_options),
                obj.get("International Note") or "",
                obj.get("Example Reference (optional)") or "",
                obj.get("Ships from") or "",
                obj.get("Handling Time (days)") or "",
                obj.get("SBAU") or "", #stock
                obj.get("Local 3PL") or "",#stock
                obj.get("China") or "",#stock
                obj.get("USA") or "",#stock
                obj.get("Main Image") or "",
                "Yes" if obj.get("Is Taxable") else "No",
            ])

        return response
