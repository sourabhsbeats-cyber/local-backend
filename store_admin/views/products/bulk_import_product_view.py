from store_admin.models.setting_model import Brand, Manufacturer
from store_admin.models.product_model import Product, ProductPriceDetails, ProductStaticAttributes, \
    ProductShippingDetails
import pandas as pd
from openpyxl import load_workbook
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from store_admin.models import Country
from store_admin.models.vendor_models import Vendor
from django.db import transaction
from django.http import FileResponse, Http404
import os
import pandas as pd
import csv
import re
from datetime import datetime
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook  # pip install openpyxl
from django.conf import settings
from django.urls import reverse
from django.utils.safestring import mark_safe
from store_admin.models.warehouse_setting_model import Warehouse
from store_admin.models.warehouse_transaction_model import ProductWarehouse
from store_admin.helpers import to_str, safe_int, safe_decimal, bool_from_str, convert_to_months

@login_required
def import_product(request):
    context = {}
    return render(request, 'sbadmin/pages/product/bulk_import/import_product_form.html', context)

def validate_field(field, value, rules):
    if rules.get("required") and (value is None or value == ""):
        return f"{field} is required."

    if value == "" or value is None:
        return None  # Skip further checks for optional fields

    if rules["type"] == "str":
        if "max" in rules and len(value) > rules["max"]:
            return f"{field} exceeds max length {rules['max']}."

    elif rules["type"] == "int":
        if not str(value).isdigit():
            return f"{field} must be an integer."

    elif rules["type"] == "bool":
        if str(value).lower() not in ["0", "1", "yes", "no", "true", "false"]:
            return f"{field} must be boolean (0/1, yes/no)."

    elif rules["type"] == "email":
        if not re.match(r"[^@]+@[^@]+\.[^@]+", value):
            return f"{field} must be a valid email."

    return None  # No errors

def validate_product_row(row_dict):
    errors = {}
    for field, rules in product_field_schema.items():
        value = row_dict.get(field)
        error = validate_field(field, value, rules)
        if error:
            errors[field] = error
    return errors

VALID_EXTENSIONS = ["csv", "xlsx"] #, "tsv", "xlsx" - FOr now csv only allowed
# "product_id": {"type": "int", "required": True},
# "created_by": {"type": "int", "required": True},
product_field_schema = {
    "Brand": {"type": "str", "required": True},
    "SKU": {"type": "str",  "required": True},
    "Title": {"type": "str", "required": True},
    "Subtitle": {"type": "str", "required": False},
    "Description": {"type": "str", "required": False},
    "Short Desc": {"type": "str", "required": False},
    "Manufacturer": {"type": "str",  "required": False},
    "EAN": {"type": "str", "required": False},
    "UPC": {"type": "str", "required": False},
    "ISBN": {"type": "str", "required": False},
    "MPN": {"type": "str", "required": False},
    "Country of Origin": {"type": "str", "required": False},
    "Condition": {"type": "str", "required": False},
    "ASIN": {"type": "str", "required": False},
    "FNSKU": {"type": "str",  "required": False},

    "FBA SKU": {"type": "str",  "required": False},
    "Is Taxable": {"type": "bool", "required": False},
    "FBA": {"type": "str", "required": False},
    "Amazon Size": {"type": "str", "required": False},
    "Barcode Label Type": {"type": "str", "required": False},
    "Prep Type": {"type": "str", "required": False},
    "Stock Status": {"type": "str", "required": False},
    "Status": {"type": "str", "required": False},
    "Publish": {"type": "str", "required": False},
    "Warranty": {"type": "str", "required": False},
    "Product Tags": {"type": "str", "required": False},

    "Sales Price": {"type": "float", "required": False},
    "Retail Price (RRP)": {"type": "float", "required": False},
    "Cost per Item": {"type": "float", "required": False},
    "Product Margin %": {"type": "float", "required": False},
    "Profit": {"type": "float", "required": False},
    "Minimum Price": {"type": "float", "required": False},
    "Maximum Price": {"type": "float", "required": False},
    "Estimated Shipping Cost": {"type": "float", "required": False},
    "Preferred Vendor": {"type": "str", "required": False},
    "Vendor SKU": {"type": "str",  "required": False},
    #"Parent SKU": {"type": "str",  "required": False},
    "Colour": {"type": "str", "required": False},
    "Size": {"type": "str", "required": False},
    "Material": {"type": "str", "required": False},
    "Compatibility": {"type": "str", "required": False},
    "Height (cm)": {"type": "float", "required": False},
    "Width (cm)": {"type": "float", "required": False},
    "Depth (cm)": {"type": "float", "required": False},
    "Weight (kg)": {"type": "float", "required": False},
    "Fast Dispatch": {"type": "str", "required": False},
    "Free Shipping": {"type": "str", "required": False},
    "Is the product bulky?": {"type": "bool", "required": False},
    #"Shipping Logic": {"type": "str", "required": False},
    "International Note": {"type": "str", "required": False},
    "Example Reference (optional)": {"type": "str", "required": False},
    "Ships from": {"type": "str", "required": False},
    "Handling Time (days)": {"type": "str", "required": False},
    #ware house
    "SBAU": {"type": "str", "max": 50, "required": False},
    "Local 3PL": {"type": "str", "required": False},
    "China": {"type": "str", "required": False},
    "USA": {"type": "str", "required": False},
    "Main Image": {"type": "str", "required": False},
}
required_headers = ['Brand', 'SKU', 'Title', 'Subtitle', 'Description', 'Short Desc',
                    'Manufacturer', 'EAN', 'UPC', 'ISBN', 'MPN', 'Country of Origin', 'Condition',
                    'ASIN', 'FNSKU', 'FBA SKU', 'FBA', 'Amazon Size', 'Barcode Label Type', 'Prep Type',
                    'Stock Status', 'Status', 'Publish', 'Warranty', 'Product Tags', 'Sales Price',
                    'Retail Price (RRP)', 'Cost per Item', 'Product Margin %', 'Profit', 'Minimum Price',
                    'Maximum Price', 'Estimated Shipping Cost', 'Preferred Vendor', 'Vendor SKU', 
                    'Colour', 'Size', 'Material', 'Compatibility', 'Height (cm)', 'Width (cm)', 'Depth (cm)',
                    'Weight (kg)', 'Fast Dispatch', 'Free Shipping', 'Is the product bulky?','Is Taxable',
                    'International Note', 'Example Reference (optional)', 'Ships from', 'Handling Time (days)',
                    'SBAU', 'Local 3PL', 'China', 'USA', 'Main Image']
cond_options = [
        {"label": "new",          "value": "New"},
        {"label": "used",         "value": "Used"},
        {"label": "refurbished",  "value": "Refurbished"},
        {"label": "open_box",     "value": "Open Box"},
    ]
barcode_options = [
    {"label": "manufacturer",   "value": "Manufacturer"},
    {"label": "amazon_barcode", "value": "Amazon Barcode"},
]
prep_type_options = [
    {"label": "no_prep_needed", "value": "No Prep Needed"},
    {"label": "polybag",        "value": "Polybagging"},
    {"label": "bubble_wrap",    "value": "Bubble Wrap"},
    {"label": "labeling",       "value": "Labeling"},
]
stock_sts_options = [
    {"label": 1, "value": "In Stock"},
    {"label": 2, "value": "Out of Stock"},
    {"label": 3, "value": "Preorder"},
    {"label": 4, "value": "Discontinued"},
]
status_options = [
    {"label": 1, "value": "Active"},
    {"label": 2, "value": "Draft"},
    {"label": 3, "value": "Discontinued"},
    {"label": 4, "value": "Archived"},
]
pub_status_options = [
    {"label": 1, "value": "Yes"},
    {"label": 0, "value": "No"},
]

#Stage 2 - Ajax upload validate - Upload file and validate
#upload the file and preview - import_product_file_upload
@login_required
def import_product_validate(request):
    if request.method != "POST":
        messages.error(request, "Invalid request.")
        return JsonResponse({"status": False, "message": "Invalid request."})

    uploaded_file = request.FILES.get("import_file")
    encoding = request.POST.get("encoding", "utf-8")
    dup = request.POST.get("dup")

    if not uploaded_file:
        return JsonResponse({"status":False, "message":"No file uploaded."})

    ext = uploaded_file.name.split(".")[-1].lower()
    if ext not in VALID_EXTENSIONS:
        return JsonResponse({"status": False, "message": "File must be CSV "})

    # Save uploaded file temporarily
    temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "product")
    os.makedirs(temp_dir, exist_ok=True)

    saved_path = os.path.join(temp_dir, uploaded_file.name)
    with open(saved_path, "wb+") as fp:
        for chunk in uploaded_file.chunks():
            fp.write(chunk)

    # Read file
    try:
        if ext == "csv":
            df = pd.read_csv(saved_path, encoding=encoding)
        elif ext == "tsv":
            df = pd.read_csv(saved_path, sep="\t", encoding=encoding)
        else:
            df = pd.read_excel(saved_path)

        file_headers = df.columns.tolist()
        #print(file_headers)
        #return JsonResponse({"status": False, "message": f"Error reading file"})
    except Exception as e:
        return JsonResponse({"status": False, "message": f"Error reading file", "error": str(e)})

    # Missing headers
    missing_headers = [h for h in required_headers if h not in file_headers]
    if missing_headers:
        msg = (
            "The uploaded file is missing required headers: "
            f"<b>{', '.join(missing_headers)}</b>"
        )
        messages.error(request, mark_safe(msg))
        os.remove(saved_path)
        return JsonResponse({"status": False, "message": "The uploaded file is missing required headers: "
            f"<b>{', '.join(missing_headers)}</b>"})
    # -------------------------
    #  CLEANING AND VALIDATION
    # -------------------------

    # Only keep mapping_fields
    cleaned_headers = list(required_headers)
    validated_rows = []
    row_errors = []

    for idx, row in df.iterrows():

        row_dict = {}
        # Build row dict ONLY from mapping fields
        for col in required_headers:
            if col in df.columns:
                val = row[col]
                row_dict[col] = "" if pd.isna(val) else str(val).strip()
            else:
                row_dict[col] = ""   # Missing field → empty

        # Validate each row
        errors = validate_product_row(row_dict)
        if errors:
            row_errors.append({
                "row": idx + 2,
                "errors": errors
            })
        else:
            validated_rows.append(row_dict)

    # If validation errors → show them
    if row_errors:
        html = "<h4>Data Errors Found:</h4><ul>"
        for row_err in row_errors:
            row_num = row_err["row"]
            err_msg = ", ".join([f"{k}: {v}" for k, v in row_err["errors"].items()])
            html += f"<li><b>Row {row_num}</b>: {err_msg}</li>"
        html += "</ul>"

        return JsonResponse({"status": False, "message": mark_safe(html)})

    # -------------------------
    #  WRITE CLEANED CSV
    # -------------------------

    cleaned_filename = f"cleaned_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    cleaned_path = os.path.join(temp_dir, cleaned_filename)

    with open(cleaned_path, "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=cleaned_headers)
        writer.writeheader()
        for row in validated_rows:
            writer.writerow({field: row.get(field, "") for field in cleaned_headers})

    # Store info in session
    # request.session["dup_option"] = dup
    # request.session["cleaned_filename"] = cleaned_filename

    # Dropdown list for mapping page

    select_options = []
    for file_header in file_headers:
        select_options.append({"value": file_header, "label": file_header})

    cleaned_filename = cleaned_filename.replace(" ", "_")
    uploaded_filename = uploaded_file.name.replace(" ", "_")
    redirect_url = reverse(
        "preview_product_import",
        kwargs={
            "cleaned_filename": cleaned_filename,
            "dup_option": request.POST.get('dub'),
            "uploaded_filename": uploaded_filename,
        }
    )
    return JsonResponse({"status": True, "message": "", "data": {
        "redirect_url": redirect_url
    }})

#Stage 3 - Confirm Import
@login_required
def preview_import(request, cleaned_filename, dup_option, uploaded_filename):
    try:
        if not cleaned_filename:
            messages.error(request, "No validated data found. Please re-upload the file.")
            return redirect("import_product")

        # Paths
        temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "product")
        cleaned_path = os.path.join(temp_dir, cleaned_filename)
        #upload_file =  os.path.join(temp_dir, uploaded_filename)
        if not os.path.exists(cleaned_path):
            messages.error(request, f"The file '{cleaned_filename}' could not be located on the server.")
            raise Exception('url_name_for_import_product_form1')

        df = pd.read_csv(cleaned_path, encoding="utf-8")
        validated_rows = df.to_dict('records')

        valid_count = len(validated_rows)

        headers = df.columns.tolist()
        # Extract headers from dict keys
        # Render preview page
        return render(request, "sbadmin/pages/product/bulk_import/import_product_stage_2.html", {
            "error": None, "dup_option":dup_option, "cleaned_filename":cleaned_filename,
            "records_preview": validated_rows[:10],  # first 10 only
            "valid_count": valid_count,  # total valid rows
            "headers": headers,
        })
    except Exception as e:
        print(str(e))
        messages.error(request, f"Error reading file: {str(e)}")
        return render(request, "sbadmin/pages/product/bulk_import/import_product_stage_2.html", {
        })

#Stage 4 - Confirm Import
@login_required
def final_product_import(request):
    if request.method != "POST":
        messages.error(request, "Invalid request.")
        return redirect("import_product")

    dup_option = request.POST.get("dup_option")  # skip_duplicate / overwrite_existing / allow_duplicates
    cleaned_filename = request.POST.get("cleaned_filename")

    if not cleaned_filename:
        messages.error(request, "No validated file. Please re-upload.")
        return redirect("import_product")

    temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "product")
    cleaned_path = os.path.join(temp_dir, cleaned_filename)

    if not os.path.exists(cleaned_path):
        messages.error(request, "Cleaned file not found. Please restart import.")
        return redirect("import_product")

    # Load cleaned CSV
    try:
        df = pd.read_csv(cleaned_path, encoding="utf-8")
    except Exception as e:
        messages.error(request, f"Could not read cleaned file: {str(e)}")
        return render(
            request,
            "sbadmin/pages/product/bulk_import/import_product_stage_3.html",
            {"imported_count": 0, "updated_count": 0},
        )

    imported_count = 0
    updated_count = 0

    # ---- Simple caches to avoid repeated DB hits ----
    brand_cache = {}
    manufacturer_cache = {}
    country_cache = {}
    vendor_cache = {}
    warehouses = list(Warehouse.objects.all())  # assume CSV has columns like "Local 3PL", "China", "USA"
    skip_count = 0
    # ---- Mapping options ----
    with transaction.atomic():
        for _, row in df.iterrows():
            sheet_sku = to_str(row.get("SKU")).strip() if row.get("SKU") else None

            # Check existing product by SKU
            existing_product = Product.objects.filter(sku=sheet_sku).first() if sheet_sku else None

            # 1) Skip duplicates → do nothing if exists
            if dup_option == "skip_duplicate" and existing_product:
                skip_count += 1
                continue

            # 2) Overwrite existing OR create new
            if dup_option == "overwrite_existing" and existing_product:
                product = existing_product
                updated_count += 1
                is_new = False
            else:
                product = Product()
                is_new = True

            # ---------- Assign SKU ----------
            if is_new:
                if sheet_sku:
                    product.sku = sheet_sku  # Use SKU from sheet if present
                else:
                    # Generate SKU only when it's a new product AND sheet SKU is missing
                    last = Product.objects.order_by("-product_id").select_for_update().first()
                    next_id = (last.product_id + 1) if last else 1
                    product.sku = f"SKU{next_id:06d}"

            # ---------- BRAND ----------
            brand_name = to_str(row.get("Brand"))
            if brand_name:
                if brand_name in brand_cache:
                    brand = brand_cache[brand_name]
                else:
                    brand = Brand.objects.filter(name=brand_name).first()
                    if not brand:
                        brand = Brand.objects.create(name=brand_name, status=1)
                    brand_cache[brand_name] = brand
                product.brand_id = brand.brand_id

            # ---------- CORE PRODUCT FIELDS ----------
            product.title = to_str(row.get("Title")) or product.title
            product.subtitle = to_str(row.get("Subtitle")) or ""
            product.description = to_str(row.get("Description")) or None
            product.short_description = to_str(row.get("Short Desc")) or None
            product.product_type = 1#standard
            # ---------- MANUFACTURER ----------
            manufacturer_name = to_str(row.get("Manufacturer"))
            if manufacturer_name:
                if manufacturer_name in manufacturer_cache:
                    manufacturer = manufacturer_cache[manufacturer_name]
                else:
                    manufacturer = Manufacturer.objects.filter(name=manufacturer_name).first()
                    if not manufacturer:
                        manufacturer = Manufacturer.objects.create(name=manufacturer_name, status=1)
                    manufacturer_cache[manufacturer_name] = manufacturer
                product.manufacturer_id = manufacturer.manufacturer_id

            # ---------- IDENTIFIERS ----------
            product.ean = to_str(row.get("EAN")) or None
            product.upc = to_str(row.get("UPC")) or None
            product.isbn = to_str(row.get("ISBN")) or None
            product.mpn = to_str(row.get("MPN")) or None

            # ---------- COUNTRY ----------
            country_name = to_str(row.get("Country of Origin"))
            if country_name:
                if country_name in country_cache:
                    country = country_cache[country_name]
                else:
                    country = Country.objects.filter(name=country_name).first()
                    country_cache[country_name] = country
                if country:
                    product.country_origin_id = country.id

            # ---------- CONDITION ----------
            cond_str = to_str(row.get("Condition"))
            if cond_str:
                condition_val = next(
                    (c["label"] for c in cond_options if c["value"].lower() == cond_str.lower()),
                    None
                )
                if condition_val:
                    product.status_condition = condition_val

            # ---------- AMAZON / FBA ----------
            product.asin = to_str(row.get("ASIN")) or None
            product.fnsku = to_str(row.get("FNSKU")) or None
            product.fba_sku = to_str(row.get("FBA SKU")) or None
            product.is_fba = bool_from_str(row.get("FBA"))
            product.is_taxable = bool_from_str(row.get("Is Taxable"))
            product.amazon_size = to_str(row.get("Amazon Size")) or ""

            # ---------- BARCODE LABEL TYPE ----------
            barcode_str = to_str(row.get("Barcode Label Type"))
            if barcode_str:
                barcode_label = next(
                    (b["label"] for b in barcode_options if b["value"].lower() == barcode_str.lower()),
                    None
                )
                if barcode_label:
                    product.barcode_label_type = barcode_label

            # ---------- PREP TYPE ----------
            prep_str = to_str(row.get("Prep Type"))
            if prep_str:
                prep_label = next(
                    (p["label"] for p in prep_type_options if p["value"].lower() == prep_str.lower()),
                    None
                )
                if prep_label:
                    product.prep_type = prep_label

            # ---------- STOCK STATUS ----------
            stock_str = to_str(row.get("Stock Status"))
            if stock_str:
                stock_label = next(
                    (s["label"] for s in stock_sts_options if s["value"].lower() == stock_str.lower()),
                    None
                )
                if stock_label is not None:
                    product.stock_status = stock_label

            # ---------- STATUS ----------
            status_str = to_str(row.get("Status"))
            if status_str:
                status_label = next(
                    (s["label"] for s in status_options if s["value"].lower() == status_str.lower()),
                    None
                )
                if status_label is not None:
                    product.status = status_label

            # ---------- PUBLISH ----------
            publish_str = to_str(row.get("Publish"))
            if publish_str:
                pub_label = next(
                    (p["label"] for p in pub_status_options if p["value"].lower() == publish_str.lower()),
                    None
                )
                if pub_label is not None:
                    product.publish_status = pub_label

            # ---------- WARRANTY / TAGS ----------
            warranty_raw = row.get("Warranty")

            product.warranty = convert_to_months(warranty_raw)
            product.product_tags = to_str(row.get("Product Tags")) or None

            # ---------- PREFERRED VENDOR ----------
            preferred_vendor_name = to_str(row.get("Preferred Vendor"))
            if preferred_vendor_name:
                if preferred_vendor_name in vendor_cache:
                    vendor = vendor_cache[preferred_vendor_name]
                else:
                    vendor = Vendor.objects.filter(display_name=preferred_vendor_name).first()

                    if not vendor:
                        # Create new vendor if not found
                        vendor = Vendor(display_name=preferred_vendor_name)

                        # Generate next unique vendor code safely
                        last = Vendor.objects.order_by("-id").select_for_update().first()
                        next_id = (last.id + 1) if last else 1

                        while Vendor.objects.filter(vendor_code=f"V{next_id:03d}").exists():
                            next_id += 1  # ensure uniqueness, only if collision happens
                        vendor.display_name = preferred_vendor_name
                        vendor.vendor_code = f"V{next_id:03d}"
                        vendor.email_address = f"{vendor.vendor_code.lower()}@tempvendor.local"
                        vendor.vendor_type = "both"
                        vendor.status = 1
                        vendor.save()

                    vendor_cache[preferred_vendor_name] = vendor

                product.preferred_vendor = vendor.id

            product.vendor_sku = to_str(row.get("Vendor SKU")) or ""
            product.parent_sku = to_str(row.get("Parent SKU")) or ""

            # created_by only for new
            if is_new:
                product.created_by = request.user.id

            # Default status if missing
            if not product.status:
                product.status = 1  # Active

            product.save()  # must save before using product_id

            # ---------- PRICE DETAILS ----------
            product_price_details = ProductPriceDetails.objects.filter(product_id=product.product_id).first()
            if not product_price_details:
                product_price_details = ProductPriceDetails(
                    product_id=product.product_id,
                    created_by=request.user.id
                )

            product_price_details.sale_price = safe_decimal(row.get("Sales Price"))
            product_price_details.retail_price = safe_decimal(row.get("Retail Price (RRP)"))
            product_price_details.cost_per_item = safe_decimal(row.get("Cost per Item"))
            product_price_details.margin_percent = safe_decimal(row.get("Product Margin %"))
            product_price_details.profit = safe_decimal(row.get("Profit"))
            product_price_details.min_price = safe_decimal(row.get("Minimum Price"))
            product_price_details.max_price = safe_decimal(row.get("Maximum Price"))
            product_price_details.estimated_shipping_cost = safe_decimal(row.get("Estimated Shipping Cost"))
            product_price_details.save()

            # ---------- STATIC ATTRIBUTES ----------
            product_static_attributes = ProductStaticAttributes.objects.filter(
                product_id=product.product_id
            ).first()
            if not product_static_attributes:
                product_static_attributes = ProductStaticAttributes(product_id=product.product_id)

            product_static_attributes.attrib_size = to_str(row.get("Size")) or ""
            product_static_attributes.attrib_color = to_str(row.get("Colour")) or ""
            product_static_attributes.attrib_material = to_str(row.get("Material")) or ""
            product_static_attributes.attrib_compatibility = to_str(row.get("Compatibility")) or ""
            product_static_attributes.attrib_height = safe_decimal(row.get("Height (cm)"))
            product_static_attributes.attrib_width = safe_decimal(row.get("Width (cm)"))
            product_static_attributes.attrib_depth = safe_decimal(row.get("Depth (cm)"))
            product_static_attributes.attrib_weight = safe_decimal(row.get("Weight (kg)"))
            product_static_attributes.save()

            # ---------- SHIPPING DETAILS ----------
            free_shipping = bool_from_str(row.get("Free Shipping"))
            bulky_str = to_str(row.get("Is the product bulky?")).lower()
            is_bulky = 1 if bulky_str == "yes" else 0 if bulky_str == "no" else None
            fast_dispatch = bool_from_str(row.get("Fast Dispatch"))

            product_shipping_details = ProductShippingDetails.objects.filter(
                product_id=product.product_id
            ).first()
            if not product_shipping_details:
                product_shipping_details = ProductShippingDetails(
                    product_id=product.product_id,
                    created_by=request.user.id
                )

            product_shipping_details.fast_dispatch = fast_dispatch
            product_shipping_details.bulky_product = is_bulky
            product_shipping_details.free_shipping = free_shipping
            product_shipping_details.international_note = to_str(row.get("International Note")) or ""
            product_shipping_details.example_reference = to_str(row.get("Example Reference (optional)")) or ""
            product_shipping_details.ships_from = to_str(row.get("Ships from")) or ""
            product_shipping_details.handling_time_days = safe_int(row.get("Handling Time (days)"), default=None)
            product_shipping_details.save()

            # ---------- WAREHOUSE STOCK ----------
            # expecting CSV columns matching warehouse.name (e.g. "Local 3PL", "China", "USA")
            for wh in warehouses:
                col_name = wh.warehouse_name
                if col_name not in df.columns:
                    continue  # column missing → ignore

                raw_stock = row.get(col_name)

                # If sheet warehouse field is EMPTY → remove mapping
                if raw_stock is None or to_str(raw_stock).strip() == "":
                    ProductWarehouse.objects.filter(
                        product_id=product.product_id,
                        warehouse_id=wh.warehouse_id
                    ).delete()
                    continue

                # Allow 0 or any integer stock value
                stock_qty = safe_int(raw_stock, default=None)
                if stock_qty is None:
                    continue  # invalid text → ignore safely

                wh_det = ProductWarehouse.objects.filter(
                    product_id=product.product_id,
                    warehouse_id=wh.warehouse_id
                ).first()

                if wh_det:
                    wh_det.stock_previous = wh_det.stock
                    wh_det.stock = stock_qty
                    wh_det.save()
                else:
                    ProductWarehouse.objects.create(
                        product_id=product.product_id,
                        warehouse_id=wh.warehouse_id,
                        stock=stock_qty,
                        stock_previous=0,
                        created_by=request.user.id,
                    )

            imported_count += 1
    os.remove(cleaned_path)
    return render(
        request,
        "sbadmin/pages/product/bulk_import/import_product_stage_3.html",
        {"imported_count": imported_count, "updated_count": updated_count},
    )

@login_required
def download_product_template(request, file_type, file_format):
    if file_type not in ["product"] or \
       file_format not in ["csv", "tsv", "xlsx"]:
        raise Http404("Invalid request")

    relative_path = f"import_templates/product/{file_type}.{file_format}"
    file_path = None
    if getattr(settings, "STATIC_ROOT", None):
        candidate = os.path.join(settings.STATIC_ROOT, relative_path)
        if os.path.exists(candidate):
            file_path = candidate

    if not file_path:
        for static_dir in getattr(settings, "STATICFILES_DIRS", []):
            candidate = os.path.join(static_dir, relative_path)
            if os.path.exists(candidate):
                file_path = candidate
                break

    if not file_path:
        raise Http404("Sample file not found")

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=f"{file_type}.{file_format}"
    )


'''
#import_product_stage_map
@login_required
def link_fields(request, cleaned_filename, dup_option, uploaded_filename):
    temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "product")
    saved_path = os.path.join(temp_dir, cleaned_filename)
    df = pd.read_csv(saved_path, encoding="utf-8")
    file_headers = df.columns.tolist()
    select_options = []
    for file_header in file_headers:
        select_options.append({"value": file_header, "label": file_header})

    cleaned_filename = cleaned_filename.replace(" ", "_")

    return render(request, "sbadmin/pages/product/bulk_import/import_product_stage_1.html", {
        "mapping_fields": required_headers,
        "select_options": select_options,
        "dup_option": dup_option,
        "cleaned_filename": cleaned_filename,
        "uploaded_filename": uploaded_filename,
    })
'''