import pandas as pd
from openpyxl import load_workbook
from openpyxl import load_workbook
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl import load_workbook  # pip install openpyxl
from openpyxl import Workbook
from rest_framework.decorators import api_view

from store_admin.helpers import getUserName
from store_admin.models import State, Country
from store_admin.models.po_models.po_models import PurchaseOrderItem, PurchaseOrder
from store_admin.models.product_model import Product

headers = ['Vendor Name', 'Vendor Code', 'Vendor SKU',
                    'PO Number', 'SBPO Order date ', 'Expected Delivery Date', 'Payment Terms', 'Warehouse',
                    'Vendor PO Number', 'Vendor PO Order Date', 'Vendor Invoice number', 'Vendor Delivery Ref#', 'Vendor Invoice Date', 'Vendor Invoice Due Date', 'Vendor Invoice Status',
                    'Shipping Provider', 'Tracking link', 'Item SKU', 'Item Name', 'Qty', 'Rate',
                    'Discount %', 'SubTotal', 'Tax %', 'Tax Total', 'Total',
                    'Freight Charge', 'Surcharges', 'Tax Total', 'Total Amount', 'Comment',
                    ]

cond_options = [
        {"label": "1", "value": "Paid"},
        {"label": "2", "value": "Unpaid"},
        {"label": "3", "value": "Cancelled"},
        {"label": "4", "value": "On Hold"},
    ]

po_order_type_options = [
    {"label": "FBM", "value": "FBM"},
    {"label": "FBA", "value": "FBA"},
    {"label": "Both", "value": "Both"},
]
ip_options = [
    {"label": 1, "value": "Yes"},
    {"label": 0, "value": "No"},
]
#use sb_commerce_b2b_dev;
export_prod_sql = """
        SELECT b.name AS "Brand",
            p.sku AS "SKU",
            p.title AS "Title",
            p.subtitle AS "Subtitle",
            p.description AS "Description",
            p.short_description AS "Short Desc",
            m.name AS "Manufacturer",
            p.ean AS "EAN",
            p.upc AS "UPC",
            p.isbn AS "ISBN",
            p.mpn AS "MPN",
            c.name  AS "Country of Origin",
            p.status_condition AS "Condition",
            p.asin AS "ASIN",
            p.fnsku AS "FNSKU",
            p.fba_sku AS "FBA SKU",
            p.is_fba AS "FBA",
            p.amazon_size AS "Amazon Size",
            p.barcode_label_type AS "Barcode Label Type",
            p.prep_type AS "Prep Type",
            p.stock_status AS "Stock Status",
            p.status AS "Status",
            p.publish_status AS "Publish",
            p.warranty AS "Warranty",
            p.product_tags AS "Product Tags",
            pp.sale_price AS "Sales Price",
            pp.retail_price AS "Retail Price (RRP)",
            pp.cost_per_item AS "Cost per Item",
            pp.margin_percent AS "Product Margin %",
            pp.profit AS "Profit",
            pp.min_price AS "Minimum Price",
            pp.max_price AS "Maximum Price",
            pp.estimated_shipping_cost AS "Estimated Shipping Cost",
            v.vendor_name  AS "Vendor Name",
            p.vendor_sku AS "Vendor SKU",
            ps.attrib_color AS "Colour",
            ps.attrib_size AS "Size",
            ps.attrib_material AS "Material",
            ps.attrib_compatibility AS "Compatibility",
            ps.attrib_height AS "Height (cm)",
            ps.attrib_width AS "Width (cm)",
            ps.attrib_depth AS "Depth (cm)",
            ps.attrib_weight AS "Weight (kg)",
            sh.fast_dispatch AS "Fast Dispatch",
            sh.free_shipping AS "Free Shipping",
            sh.bulky_product AS "Is the product bulky?",
            sh.international_note AS "International Note",
            sh.example_reference AS "Example Reference (optional)",
            sh.ships_from AS "Ships from",
            sh.handling_time_days AS "Handling Time (days)",
            /*COALESCE(
               (SELECT SUM(stock)
                FROM store_admin_product_warehouse
                WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'China')
                      AND product_id = p.product_id
               ), 0
            ) AS "China",
        
            COALESCE(
               (SELECT SUM(stock)
                FROM store_admin_product_warehouse
                WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'Local 3PL')
                      AND product_id = p.product_id
               ), 0
            ) AS "Local 3PL",
        
            COALESCE(
               (SELECT SUM(stock)
                FROM store_admin_product_warehouse
                WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'SBAU')
                      AND product_id = p.product_id
               ), 0
            ) AS "SBAU",
        
            COALESCE(
               (SELECT SUM(stock)
                FROM store_admin_product_warehouse
                WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'USA')
                      AND product_id = p.product_id
               ), 0
            ) AS "USA",
            p.main_image AS "Main Image",*/
            p.is_taxable AS "Is Taxable"
            FROM store_admin_product p
             LEFT JOIN store_admin_brand b ON b.brand_id = p.brand_id
             LEFT JOIN store_admin_manufacturer m ON m.manufacturer_id = p.manufacturer_id
             LEFT JOIN store_admin_countries c ON c.id = p.country_origin_id
             LEFT JOIN store_admin_vendor v ON v.id = p.preferred_vendor
             LEFT JOIN store_admin_product_price_details pp ON pp.product_id = p.product_id
             LEFT JOIN store_admin_product_static_attributes ps ON ps.product_id = p.product_id
             LEFT JOIN store_admin_product_shipping_details sh ON sh.product_id = p.product_id
             LEFT JOIN store_admin_product_warehouse pw1 ON pw1.product_id = p.product_id AND pw1.warehouse_id = 1
             LEFT JOIN store_admin_product_warehouse pw2 ON pw2.product_id = p.product_id AND pw2.warehouse_id = 2
             LEFT JOIN store_admin_product_warehouse pw3 ON pw3.product_id = p.product_id AND pw3.warehouse_id = 3
        ORDER BY p.product_id ASC
                  
                  """
def map_option(value, options):
    if value is None or value == "":
        return ""
        # If input is integer → match directly without lower()
    if isinstance(value, int):
        return next((o["value"] for o in options if o["value"] == value), "")
        # else process as string
    v = str(value).strip().lower()
    return next((o["value"] for o in options if o["value"].lower() == v), "")


def map_int_option(val, options):
    if val is None:   # handles NULL safely
        return ""
    return next((o["value"] for o in options if o["label"] == val), "")

def map_label_option(val, options):
    # handle null/empty safely for CSV
    if val is None or val == "":
        return ""

    # int input → match label directly
    if isinstance(val, int):
        return next((o["value"] for o in options if o["label"] == val), "")

    # string input → compare with option.label case-insensitive
    v = str(val).strip().lower()
    return next(
        (o["value"] for o in options if str(o["label"]).strip().lower() == v),
        ""
    )

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import connection
from datetime import datetime
from django.utils import timezone
from decimal import Decimal, ROUND_HALF_UP
@api_view(['GET'])
def export_po_orders(request):
    timestamp = timezone.now().strftime("%d%m%Y_%H%M%S")
    file_name = f"po_export_{timestamp}.csv"

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{file_name}"'
    response["Access-Control-Expose-Headers"] = "Content-Disposition"

    writer = csv.writer(response)

    # ---------- HEADERS ----------
    writer.writerow([
        "SBPO Number",
        "Vendor Name",
        "Vendor Code",
        "Minimum Purchase",
        "Discount (%)",
        "GST Rate (%)",
        "SB Warehouse",
        "Delivery Name",
        "Address Line 1",
        "Address Line 2",
        "Country",
        "State",
        "City",
        "ZIP",
        "SBPO Order Date",
        "Delivery Date",
        "Item SKU",
        "Item Name",
        "ASIN",
        "Qty",
        "Product Rate",
        "Product Discount %",
        "Product GST %",
        "Product GST Amount",
        "Sub Total",
        "Landed Cost",
        "Line Total",
        "Line Comment",
        "Subtotal",
        "Surcharge",
        "Freight",
        "GST Total",
        "Grand Total",
        "Comment",
        "Created At",
        "Created By",
    ])

    purchase_orders = PurchaseOrder.objects.order_by("po_id")

    for po in purchase_orders:

        state_obj = State.objects.filter(id=po.state).first() if po.state else None
        country_obj = Country.objects.filter(id=po.country_id).first() if po.country_id else None

        order_date_str = po.order_date.strftime("%d/%m/%Y") if po.order_date else ""
        delivery_date_str = po.delivery_date.strftime("%d/%m/%Y") if po.delivery_date else ""
        created_at_str = po.created_at.strftime("%d/%m/%Y %I:%M %p") if po.created_at else ""
        created_by_name = getUserName(po.created_by) if po.created_by else ""

        items = list(PurchaseOrderItem.objects.filter(po_id=po.po_id))
        if not items:
            continue

        # ---------- Decimal Safe Total Qty ----------
        total_qty = sum(Decimal(str(i.qty or 0)) for i in items)

        # ---------- Decimal Safe Additional Per Unit ----------
        additional_per_unit = Decimal("0")
        if total_qty > 0:
            shipping = Decimal(str(po.shipping_charge or 0))
            surcharge = Decimal(str(po.surcharge_total or 0))
            additional_per_unit = (shipping + surcharge) / total_qty

        # ---------- Optimize Product Fetch ----------
        product_ids = [i.product_id for i in items]
        products = {
            p.product_id: p
            for p in Product.objects.filter(product_id__in=product_ids)
        }

        for item in items:
            product = products.get(item.product_id)

            qty = Decimal(str(item.qty or 0))
            price = Decimal(str(item.price or 0))
            discount = Decimal(str(item.discount_percentage or 0))

            discounted_unit_price = price * (
                Decimal("1") - (discount / Decimal("100"))
            )

            landed_cost = discounted_unit_price + additional_per_unit
            landed_cost = landed_cost.quantize(
                Decimal("0.0001"), rounding=ROUND_HALF_UP
            )

            writer.writerow([
                po.po_number or "",
                po.vendor_name or "",
                po.vendor_code or "",
                po.minimum_order_value or 0,
                po.global_discount_percentage or 0,
                po.global_tax_rate or 0,

                po.warehouse_id or "",
                po.delivery_name or "",
                po.address_line1 or "",
                po.address_line2 or "",
                country_obj.name if country_obj else "",
                state_obj.name if state_obj else "",
                po.city or "",
                po.post_code or "",

                order_date_str,
                delivery_date_str,

                product.sku if product else "",
                product.title if product else "",
                product.asin if product else "",

                qty,
                price,
                discount,
                item.tax_percentage or 0,
                item.tax_amount or 0,
                item.subtotal or 0,
                landed_cost,
                item.line_total or 0,
                item.comment or "",

                po.sub_total or 0,
                po.surcharge_total or 0,
                po.shipping_charge or 0,
                po.tax_total or 0,
                po.summary_total or 0,

                po.comments or "",
                created_at_str,
                created_by_name
            ])

    return response

