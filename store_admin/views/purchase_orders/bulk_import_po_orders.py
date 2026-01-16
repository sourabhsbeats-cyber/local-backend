from django.utils import timezone

from store_admin.models.po_models.po_models import PurchaseOrderVendor, PurchaseOrder, POStatus, PurchaseOrderItem
from store_admin.models.setting_model import Brand, Manufacturer
from store_admin.models.product_model import Product, ProductPriceDetails, ProductStaticAttributes, \
    ProductShippingDetails
import pandas as pd
from openpyxl import load_workbook
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import JsonResponse
from store_admin.models import Country, PaymentTerm
from store_admin.models.vendor_models import Vendor
from django.db import transaction
from django.http import FileResponse, Http404
import os
import pandas as pd
import csv
import re
from datetime import datetime
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
from openpyxl import load_workbook  # pip install openpyxl
from django.conf import settings
from django.urls import reverse
from django.utils.safestring import mark_safe
from store_admin.models.warehouse_setting_model import Warehouse
from store_admin.models.warehouse_transaction_model import ProductWarehouse
from store_admin.helpers import to_str, safe_int, safe_decimal, bool_from_str, convert_to_months, parse_date_or_none


@login_required
def import_purchase_order(request):
    context = {}
    return render(request, 'sbadmin/pages/purchase_order/bulk_import/import_po_form.html', context)

def validate_field(field, value, rules):
    if rules.get("required") and (value is None or value == ""):
        return f"{field} is required."

    if value == "" or value is None:
        return None  # Skip further checks for optional fields

    if rules["type"] == "str":
        if "max" in rules and len(value) > rules["max"]:
            return f"{field} exceeds max length {rules['max']}."

    elif rules["type"] == "int":
        if not str(value).isdigit():
            return f"{field} must be an integer."

    elif rules["type"] == "bool":
        if str(value).lower() not in ["0", "1", "yes", "no", "true", "false"]:
            return f"{field} must be boolean (0/1, yes/no)."

    elif rules["type"] == "email":
        if not re.match(r"[^@]+@[^@]+\.[^@]+", value):
            return f"{field} must be a valid email."

    return None  # No errors

def validate_product_row(row_dict):
    errors = {}
    for field, rules in po_field_schema.items():
        value = row_dict.get(field)
        error = validate_field(field, value, rules)
        if error:
            errors[field] = error
    return errors

VALID_EXTENSIONS = ["csv", "xlsx"] #, "tsv", "xlsx" - FOr now csv only allowed
# "product_id": {"type": "int", "required": True},
# "created_by": {"type": "int", "required": True},
po_field_schema = {
    "Vendor Name": {"type": "str", "required": False},
    "Vendor Code": {"type": "str",  "required": True},
    #"Vendor SKU": {"type": "str", "required": False},
    "PO Number": {"type": "str", "required": False},
    "SBPO Order date": {"type": "str", "required": True},
    "Expected Delivery Date": {"type": "str", "required": True},
    "Payment Terms": {"type": "str",  "required": False},
    "Warehouse": {"type": "str", "required": False},
    "Vendor PO Number": {"type": "str", "required": False},

    "Vendor PO Order Date": {"type": "str", "required": False},
    "Vendor Invoice number": {"type": "str", "required": False},
    "Vendor Delivery Ref#": {"type": "str", "required": False},
    "Vendor Invoice Date": {"type": "str", "required": False},
    "Vendor Invoice Due Date": {"type": "str", "required": False},
    "Vendor Invoice Status": {"type": "str",  "required": False},

    "Shipping Provider": {"type": "str",  "required": False},
    "Tracking link": {"type": "bool", "required": False},

    "Item SKU": {"type": "str", "required": True},
    "Item Name": {"type": "str", "required": False},
    "Qty": {"type": "str", "required": True},
    "Rate": {"type": "str", "required": True},
    "Discount %": {"type": "str", "required": True},
    "Tax %": {"type": "str", "required": False},
    "Tax Amount": {"type": "str", "required": False},
    "SubTotal": {"type": "str", "required": False},
   # "Total": {"type": "str", "required": False},
    "Freight Charge": {"type": "float", "required": False},
    "Surcharges": {"type": "float", "required": False},
    #"Tax Total": {"type": "float", "required": False},
    #"Total Amount": {"type": "float", "required": False},
    "Comment": {"type": "float", "required": False},
}
po_import_headers = ['Vendor Name', 'Vendor Code','PO Number', 'SBPO Order date', 'Expected Delivery Date',
            'Payment Terms', 'Warehouse', 'Vendor PO Number', 'Vendor PO Order Date', 'Vendor Invoice number',
            'Vendor Delivery Ref#', 'Vendor Invoice Date', 'Vendor Invoice Due Date', 'Vendor Invoice Status',
           'Shipping Provider', 'Tracking link', 'Item SKU', 'Item Name', 'Qty', 'Rate',
           'Discount %', 'Tax %', 'Tax Amount', 'SubTotal',
           'Freight Charge', 'Surcharges', 'Comment',
        ]


inv_payment_options = [
        {"label": 1, "value": "Paid"},
        {"label": 2, "value": "Unpaid"},
        {"label": 3, "value": "Cancelled"},
        {"label": 4, "value": "On Hold"},
    ]

po_order_type_options = [
    {"label": "FBM",   "value": "FBM"},
    {"label": "FBA", "value": "FBA"},
    {"label": "Both", "value": "Both"},
]
#Stage 2 - Ajax upload validate - Upload file and validate
#upload the file and preview - import_product_file_upload

from decimal import Decimal, ROUND_HALF_UP
import math
def d(val):
    if val is None:
        return Decimal("0")

    # pandas NaN / float nan
    if isinstance(val, float) and math.isnan(val):
        return Decimal("0")

    val = str(val).strip()

    # string NaN
    if val.lower() == "nan" or val == "":
        return Decimal("0")

    # CLEAN CURRENCY + COMMAS
    val = val.replace("$", "").replace(",", "")

    try:
        return Decimal(val)
    except Exception:
        return Decimal("0")

def q(val):
    return val.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def clean_decimal(val):
    if val is None or pd.isna(val):
        return Decimal("0")
    val = str(val).strip()
    if val == "":
        return Decimal("0")
    val = val.replace("$", "").replace(",", "")
    return Decimal(val)


@login_required
def import_po_validate(request):
    if request.method != "POST":
        messages.error(request, "Invalid request.")
        return JsonResponse({"status": False, "message": "Invalid request."})

    uploaded_file = request.FILES.get("import_file")
    encoding = request.POST.get("encoding", "utf-8")

    if not uploaded_file:
        return JsonResponse({"status": False, "message": "No file uploaded."})

    ext = uploaded_file.name.split(".")[-1].lower()
    if ext not in VALID_EXTENSIONS:
        return JsonResponse({"status": False, "message": "File must be CSV"})

    # -------------------------
    # SAVE TEMP FILE
    # -------------------------
    temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "purchase_orders")
    os.makedirs(temp_dir, exist_ok=True)

    saved_path = os.path.join(temp_dir, uploaded_file.name)
    with open(saved_path, "wb+") as fp:
        for chunk in uploaded_file.chunks():
            fp.write(chunk)

    # -------------------------
    # READ FILE
    # -------------------------
    try:
        if ext == "csv":
            df = pd.read_csv(saved_path, encoding=encoding)
        elif ext == "tsv":
            df = pd.read_csv(saved_path, sep="\t", encoding=encoding)
        else:
            df = pd.read_excel(saved_path)

        if df.empty:
            return JsonResponse({"status": False, "message": "Empty rows"})
    except Exception as e:
        return JsonResponse({"status": False, "message": "Error reading file", "error": str(e)})

    file_headers = df.columns.tolist()

    # -------------------------
    # HEADER VALIDATION
    # -------------------------
    missing_headers = [h for h in po_field_schema if h not in file_headers]
    if missing_headers:
        msg = "Missing required headers: <b>%s</b>" % ", ".join(missing_headers)
        messages.error(request, mark_safe(msg))
        return JsonResponse({"status": False, "message": mark_safe(msg)})

    # -------------------------
    # ROW VALIDATION
    # -------------------------
    validated_rows = []
    row_errors = []

    for idx, row in df.iterrows():
        row_dict = {}

        for col in po_field_schema:
            val = row[col] if col in df.columns else ""
            # Special handling for numeric money fields
            if col in ("Freight Charge", "Surcharges"):
                if pd.isna(val) or str(val).strip() == "":
                    row_dict[col] = Decimal("0")
                else:
                    row_dict[col] = Decimal(str(val))
            else:
                row_dict[col] = "" if pd.isna(val) else str(val).strip()

        errors = validate_product_row(row_dict)
        if errors:
            row_errors.append({"row": idx + 2, "errors": errors})
        else:
            validated_rows.append(row_dict)

    if row_errors:
        html = "<h4>Data Errors Found:</h4><ul>"
        for err in row_errors:
            msgs = ", ".join(f"{k}: {v}" for k, v in err["errors"].items())
            html += f"<li><b>Row {err['row']}</b>: {msgs}</li>"
        html += "</ul>"
        return JsonResponse({"status": False, "message": mark_safe(html)})

    # -------------------------
    # SORT & CALCULATE
    # -------------------------
    validated_rows = sorted(
        validated_rows,
        key=lambda x: (x.get("PO Number", ""), x.get("Item SKU", ""))
    )

    for row in validated_rows:
        qty = d(row.get("Qty"))
        rate = d(row.get("Rate"))
        discount_pct = d(row.get("Discount %"))
        tax_raw = row.get("Tax %")

        # Base
        base_raw = qty * rate

        # Tax
        if tax_raw == "" or tax_raw is None:
            tax_pct = Decimal("10")
        else:
            tax_pct = d(tax_raw)

        tax_amount = base_raw * tax_pct / Decimal("100")
        subtotal = base_raw + tax_amount

        # Discount (after tax)
        if discount_pct > 0:
            subtotal -= subtotal * discount_pct / Decimal("100")

        row["Tax Amount"] = str(q(tax_amount))
        row["SubTotal"] = str(q(subtotal))

    # -------------------------
    # WRITE CLEANED CSV
    # -------------------------
    cleaned_filename = f"cleaned_{datetime.now().strftime('%Y%m%d%H%M%S')}.csv"
    cleaned_path = os.path.join(temp_dir, cleaned_filename)

    with open(cleaned_path, "w", newline="", encoding="utf-8") as fp:
        writer = csv.DictWriter(fp, fieldnames=po_field_schema)
        writer.writeheader()
        for row in validated_rows:
            writer.writerow({h: row.get(h, "") for h in po_field_schema})

    # -------------------------
    # RESPONSE
    # -------------------------
    redirect_url = reverse(
        "preview_po_import",
        kwargs={
            "cleaned_filename": cleaned_filename.replace(" ", "_"),
            "dup_option": request.POST.get("dup"),
            "uploaded_filename": uploaded_file.name.replace(" ", "_"),
        }
    )

    return JsonResponse({"status": True, "data": {"redirect_url": redirect_url}})

def first_non_empty(rows, column):
    """
    Iterate all rows of same PO
    Return first non-empty value for given column
    """
    for r in rows:
        val = r.get(column)
        if val is None:
            continue
        val = str(val).strip()
        if val != "":
            return val
    return None


#Stage 3 - Confirm Import
@login_required
def preview_import(request, cleaned_filename, dup_option, uploaded_filename):
    try:
        if not cleaned_filename:
            messages.error(request, "No validated data found. Please re-upload the file.")
            return redirect("import_product")

        # Paths
        temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "purchase_orders")
        cleaned_path = os.path.join(temp_dir, cleaned_filename)
        #upload_file =  os.path.join(temp_dir, uploaded_filename)
        if not os.path.exists(cleaned_path):
            messages.error(request, f"The file '{cleaned_filename}' could not be located on the server.")
            raise Exception('url_name_for_import_product_form1')

        df = pd.read_csv(cleaned_path, encoding="utf-8")
        validated_rows = df.to_dict('records')

        valid_count = len(validated_rows)

        headers = df.columns.tolist()
        # Extract headers from dict keys
        # Render preview page
        return render(request, "sbadmin/pages/purchase_order/bulk_import/import_po_stage_2.html", {
            "error": None, "dup_option":dup_option, "cleaned_filename":cleaned_filename,
            "records_preview": validated_rows[:10],  # first 10 only
            "valid_count": valid_count,  # total valid rows
            "headers": headers,
        })
    except Exception as e:
        print(str(e))
        messages.error(request, f"Error reading file: {str(e)}")
        return render(request, "sbadmin/pages/purchase_order/bulk_import/import_po_stage_2.html", {
        })

def set_if_empty(target_dict, key, value):
    if not target_dict.get(key) and value not in ("", None):
        target_dict[key] = value

@login_required
def final_po_import(request):
    if request.method != "POST":
        messages.error(request, "Invalid request.")
        return redirect("import_po_orders")

    cleaned_filename = request.POST.get("cleaned_filename")
    if not cleaned_filename:
        messages.error(request, "No validated file found.")
        return redirect("import_po_orders")

    temp_dir = os.path.join(settings.MEDIA_ROOT, "imports", "purchase_orders")
    cleaned_path = os.path.join(temp_dir, cleaned_filename)

    if not os.path.exists(cleaned_path):
        messages.error(request, "Cleaned file not found.")
        return redirect("import_po_orders")

    # -------------------------
    # READ CSV
    # -------------------------
    try:
        df = pd.read_csv(cleaned_path, encoding="utf-8")
    except Exception as e:
        messages.error(request, f"Error reading file: {e}")
        return redirect("import_po_orders")

    if df.empty:
        messages.error(request, "Empty file.")
        return redirect("import_po_orders")

    # Ensure correct processing order
    df = df.sort_values("PO Number")

    imported_count = 0

    # -------------------------
    # GROUPING VARIABLES
    # -------------------------
    current_po_number = None
    current_po = None

    po_sub_total = Decimal("0")
    po_tax_total = Decimal("0")
    po_freight = Decimal("0")
    po_surcharge = Decimal("0")

    vendor_values = {}
    with transaction.atomic():

        for _, row in df.iterrows():
            GLOBAL_TAX_PCT = d(row.get("Tax %"))
            po_number = str(row.get("PO Number", "")).strip()

            # -------------------------
            # PO CHANGE DETECTION
            # -------------------------
            if po_number != current_po_number:

                # SAVE PREVIOUS PO TOTALS
                if current_po:
                    current_po.sub_total = po_sub_total
                    current_po.tax_total = po_tax_total
                    current_po.shipping_charge = po_freight
                    current_po.surcharge_total = po_surcharge
                    current_po.save(update_fields=[
                        "sub_total",
                        "tax_total",
                        "shipping_charge",
                        "surcharge_total",
                    ])

                # RESET TOTALS
                po_sub_total = Decimal("0")
                po_tax_total = Decimal("0")
                po_freight = d(row.get("Freight Charge"))
                po_surcharge = d(row.get("Surcharges"))

                #  CHECK IF PO ALREADY EXISTS
                if PurchaseOrder.objects.filter(
                        po_number=po_number,
                        is_archived=0
                ).exists():
                    current_po = None
                    current_po_number = po_number
                    continue  #  skip ALL rows of this PO

                # FETCH VENDOR
                vendor_code = str(row.get("Vendor Code", "")).strip()
                vendor = Vendor.objects.filter(vendor_code=vendor_code).first()
                if not vendor:
                    raise Exception(f"Vendor not found: {vendor_code}")

                payment_term_name = str(row.get("Payment Terms", "")).strip()
                payment_term = (
                    PaymentTerm.objects.filter(name__icontains=payment_term_name).first()
                    if payment_term_name else None
                )

                # CREATE NEW PO
                current_po = PurchaseOrder.objects.create(
                    po_number=po_number,
                    vendor_id=vendor.id,
                    vendor_code=vendor.vendor_code,
                    vendor_name=vendor.display_name,
                    currency_code=vendor.currency,
                    tax_percentage=GLOBAL_TAX_PCT,
                    payment_term_id=payment_term.id if payment_term else None,
                    order_date=parse_date_or_none(row.get("SBPO Order date")),
                    delivery_date=parse_date_or_none(row.get("Expected Delivery Date")),
                    shipping_charge=po_freight,
                    surcharge_total=po_surcharge,
                    status_id=POStatus.PARKED,
                    created_by=request.user.id,
                    created_at=timezone.now(),
                    comments=str(row.get("Comment"))
                )

                # -------------------------
                # CREATE PURCHASE ORDER VENDOR (ONE PER PO)
                # -------------------------

                current_po_number = po_number
                imported_count += 1

            # -------------------------
            # COLLECT VENDOR VALUES (SPREAD ACROSS ROWS)
            # -------------------------
            set_if_empty(vendor_values, "vendor_po_number",
                         row.get("Vendor PO Number"))

            set_if_empty(vendor_values, "vendor_invoice_number",
                         row.get("Vendor Invoice number"))

            set_if_empty(vendor_values, "vendor_delivery_ref",
                         row.get("Vendor Delivery Ref#"))

            po_order_date = parse_date_or_none(row.get("Vendor PO Order Date"))
            if po_order_date:
                vendor_values.setdefault("vendor_po_order_date", po_order_date)

            invoice_date = parse_date_or_none(row.get("Vendor Invoice Date"))
            if invoice_date:
                vendor_values.setdefault("vendor_invoice_date", invoice_date)

            invoice_due_date = parse_date_or_none(row.get("Vendor Invoice Due Date"))
            if invoice_due_date:
                vendor_values.setdefault("vendor_invoice_due_date", invoice_due_date)

            set_if_empty(vendor_values, "vendor_invoice_status",
                         row.get("Vendor Invoice Status"))

            # -------------------------
            # CREATE ITEM ONLY IF PO CREATED
            # -------------------------
            if not current_po:
                continue

            product_sku = str(row.get("Item SKU", "")).strip()
            product = Product.objects.filter(sku=product_sku).first()
            if not product:
                raise Exception(f"Product not found: {product_sku}")

            item = PurchaseOrderItem.objects.create(
                po_id=current_po.po_id,
                product_id=product.product_id,
                qty=int(row.get("Qty", 0)),
                price=d(row.get("Rate")),
                tax_percentage=GLOBAL_TAX_PCT,
                discount_percentage=d(row.get("Discount %")),
                created_by=request.user.id,
                created_at=timezone.now(),
            )

            po_sub_total += item.subtotal
            po_tax_total += item.tax_amount

            # -------------------------
            # SAVE LAST PO TOTALS
            # -------------------------
        if current_po:
            if vendor_values:
                PurchaseOrderVendor.objects.create(
                    po_id=current_po.po_id,
                    created_by=request.user.id,
                    po_number=vendor_values.get("vendor_po_number"),
                    invoice_ref_number=vendor_values.get("vendor_invoice_number"),
                    delivery_ref_number=vendor_values.get("vendor_delivery_ref"),
                    #invoice_status=vendor_values.get("vendor_invoice_status"),
                    invoice_due_date=vendor_values.get("vendor_invoice_due_date"),
                    order_date=vendor_values.get("vendor_po_order_date"),
                    invoice_date=vendor_values.get("vendor_invoice_date")
                )

            current_po.sub_total = po_sub_total
            current_po.tax_total = po_tax_total
            current_po.shipping_charge = po_freight
            current_po.surcharge_total = po_surcharge
            current_po.save(update_fields=[
                "sub_total",
                "tax_total",
                "shipping_charge",
                "surcharge_total",
            ])

    messages.success(request, f"{imported_count} Purchase Orders imported successfully.")
   # return redirect("purchase_order_list")

    os.remove(cleaned_path)
    return render(
        request,
        "sbadmin/pages/purchase_order/bulk_import/import_po_stage_3.html",
        {"imported_count": imported_count, "updated_count": "--"},
    )

@login_required
def download_po_template(request, file_type, file_format):
    if file_type not in ["po_template"] or \
       file_format not in ["csv", "tsv", "xlsx"]:
        raise Http404("Invalid request")

    relative_path = f"import_templates/purchase_orders/{file_type}.{file_format}"
    file_path = None
    if getattr(settings, "STATIC_ROOT", None):
        candidate = os.path.join(settings.STATIC_ROOT, relative_path)
        if os.path.exists(candidate):
            file_path = candidate

    if not file_path:
        for static_dir in getattr(settings, "STATICFILES_DIRS", []):
            candidate = os.path.join(static_dir, relative_path)
            if os.path.exists(candidate):
                file_path = candidate
                break

    if not file_path:
        raise Http404("Sample file not found")

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename=f"{file_type}.{file_format}"
    )
