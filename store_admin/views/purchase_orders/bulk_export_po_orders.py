import pandas as pd
from openpyxl import load_workbook
from openpyxl import load_workbook
import pandas as pd
import csv
from openpyxl import load_workbook
from openpyxl import load_workbook  # pip install openpyxl
from openpyxl import Workbook

from store_admin.models.po_models.po_models import PurchaseOrderItem, PurchaseOrder

headers = ['Vendor Name', 'Vendor Code', 'Vendor SKU',
                    'PO Number', 'SBPO Order date ', 'Expected Delivery Date', 'Payment Terms', 'Warehouse',
                    'Vendor PO Number', 'Vendor PO Order Date', 'Vendor Invoice number', 'Vendor Delivery Ref#', 'Vendor Invoice Date', 'Vendor Invoice Due Date', 'Vendor Invoice Status',
                    'Shipping Provider', 'Tracking link', 'Item SKU', 'Item Name', 'Qty', 'Rate',
                    'Discount %', 'SubTotal', 'Tax %', 'Tax Total', 'Total',
                    'Freight Charge', 'Surcharges', 'Tax Total', 'Total Amount', 'Comment',
                    ]

cond_options = [
        {"label": "1",          "value": "Paid"},
        {"label": "2",         "value": "Unpaid"},
        {"label": "3",  "value": "Cancelled"},
        {"label": "4",     "value": "On Hold"},
    ]

po_order_type_options = [
    {"label": "FBM",   "value": "FBM"},
    {"label": "FBA", "value": "FBA"},
    {"label": "Both", "value": "Both"},
]
ip_options = [
    {"label": 1, "value": "Yes"},
    {"label": 0, "value": "No"},
]
export_prod_sql = """
                   SELECT b.name AS "Brand",
	p.sku AS "SKU",
	p.title AS "Title",
	p.subtitle AS "Subtitle",
	p.description AS "Description",
	p.short_description AS "Short Desc",
	m.name AS "Manufacturer",
	p.ean AS "EAN",
	p.upc AS "UPC",
	p.isbn AS "ISBN",
	p.mpn AS "MPN",
	c.name  AS "Country of Origin",
	p.status_condition AS "Condition",
	p.asin AS "ASIN",
	p.fnsku AS "FNSKU",
	p.fba_sku AS "FBA SKU",
	p.is_fba AS "FBA",
	p.amazon_size AS "Amazon Size",
	p.barcode_label_type AS "Barcode Label Type",
	p.prep_type AS "Prep Type",
	p.stock_status AS "Stock Status",
	p.status AS "Status",
	p.publish_status AS "Publish",
	p.warranty AS "Warranty",
	p.product_tags AS "Product Tags",
	pp.sale_price AS "Sales Price",
	pp.retail_price AS "Retail Price (RRP)",
	pp.cost_per_item AS "Cost per Item",
	pp.margin_percent AS "Product Margin %",
	pp.profit AS "Profit",
	pp.min_price AS "Minimum Price",
	pp.max_price AS "Maximum Price",
	pp.estimated_shipping_cost AS "Estimated Shipping Cost",
	v.display_name  AS "Preferred Vendor",
	p.vendor_sku AS "Vendor SKU",
	ps.attrib_color AS "Colour",
	ps.attrib_size AS "Size",
	ps.attrib_material AS "Material",
	ps.attrib_compatibility AS "Compatibility",
	ps.attrib_height AS "Height (cm)",
	ps.attrib_width AS "Width (cm)",
	ps.attrib_depth AS "Depth (cm)",
	ps.attrib_weight AS "Weight (kg)",
	sh.fast_dispatch AS "Fast Dispatch",
	sh.free_shipping AS "Free Shipping",
	sh.bulky_product AS "Is the product bulky?",
	sh.international_note AS "International Note",
	sh.example_reference AS "Example Reference (optional)",
	sh.ships_from AS "Ships from",
	sh.handling_time_days AS "Handling Time (days)",
	COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'China')
              AND product_id = p.product_id
       ), 0
    ) AS "China",

    COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'Local 3PL')
              AND product_id = p.product_id
       ), 0
    ) AS "Local 3PL",

    COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'SBAU')
              AND product_id = p.product_id
       ), 0
    ) AS "SBAU",

    COALESCE(
       (SELECT SUM(stock)
        FROM store_admin_product_warehouse
        WHERE warehouse_id = (SELECT warehouse_id FROM store_admin_warehouse WHERE warehouse_name = 'USA')
              AND product_id = p.product_id
       ), 0
    ) AS "USA",
    p.main_image AS "Main Image",
    p.is_taxable AS "Is Taxable"
	FROM store_admin_product p
	 LEFT JOIN store_admin_brand b ON b.brand_id = p.brand_id
	 LEFT JOIN store_admin_manufacturer m ON m.manufacturer_id = p.manufacturer_id
	 LEFT JOIN store_admin_countries c ON c.id = p.country_origin_id
	 LEFT JOIN store_admin_vendor v ON v.id = p.preferred_vendor
	 LEFT JOIN store_admin_product_price_details pp ON pp.product_id = p.product_id
	 LEFT JOIN store_admin_product_static_attributes ps ON ps.product_id = p.product_id
	 LEFT JOIN store_admin_product_shipping_details sh ON sh.product_id = p.product_id
	 LEFT JOIN store_admin_product_warehouse pw1 ON pw1.product_id = p.product_id AND pw1.warehouse_id = 1
	 LEFT JOIN store_admin_product_warehouse pw2 ON pw2.product_id = p.product_id AND pw2.warehouse_id = 2
	 LEFT JOIN store_admin_product_warehouse pw3 ON pw3.product_id = p.product_id AND pw3.warehouse_id = 3
ORDER BY p.product_id ASC
                    """
def map_option(value, options):
    if value is None or value == "":
        return ""
        # If input is integer → match directly without lower()
    if isinstance(value, int):
        return next((o["value"] for o in options if o["value"] == value), "")
        # else process as string
    v = str(value).strip().lower()
    return next((o["value"] for o in options if o["value"].lower() == v), "")


def map_int_option(val, options):
    if val is None:   # ✅ handles NULL safely
        return ""
    return next((o["value"] for o in options if o["label"] == val), "")

def map_label_option(val, options):
    # handle null/empty safely for CSV
    if val is None or val == "":
        return ""

    # int input → match label directly
    if isinstance(val, int):
        return next((o["value"] for o in options if o["label"] == val), "")

    # string input → compare with option.label case-insensitive
    v = str(val).strip().lower()
    return next(
        (o["value"] for o in options if str(o["label"]).strip().lower() == v),
        ""
    )

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db import connection

@login_required
def export_po_orders_xlsx(request):

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="po_reverse_export.csv"'

    writer = csv.writer(response)

    # SAME HEADERS AS IMPORT
    writer.writerow([
        "Vendor Name",
        "Vendor Code",
        "PO Number",
        "SBPO Order date",
        "Expected Delivery Date",
        "Payment Terms",
        "Warehouse",
        "Vendor PO Number",
        "Vendor PO Order Date",
        "Vendor Invoice number",
        "Vendor Delivery Ref#",
        "Vendor Invoice Date",
        "Vendor Invoice Due Date",
        "Vendor Invoice Status",
        "Shipping Provider",
        "Tracking link",
        "Item SKU",
        "Item Name",
        "Qty",
        "Rate",
        "Discount %",
        "Tax %",
        "Tax Amount",
        "SubTotal",
        "Freight Charge",
        "Surcharges",
        "Comment",
    ])

    # 🔥 STEP 1: GET ALL POs
    purchase_orders = PurchaseOrder.objects.order_by("po_id")

    for po in purchase_orders:

        # STEP 2: COLLECT PO COMMON VALUES
        po_common = {
            "vendor_name": po.vendor_name or "",
            "vendor_code": po.vendor_code or "",
            "po_number": po.po_number or "",
            "order_date": po.order_date or "",
            "delivery_date": po.delivery_date or "",
            "payment_term": po.payment_term_id or "",
            "warehouse": po.warehouse_id or "",
            "vendor_po_number": po.vendor_reference or "",
            "freight": po.shipping_charge or 0,
            "surcharge": po.surcharge_total or 0,
            "comment": po.comments or "",
        }

        # STEP 3: GET ITEMS FOR THIS PO
        items = PurchaseOrderItem.objects.filter(po_id=po.po_id)

        # STEP 4: LOOP ITEMS → WRITE ROWS
        for item in items:
            writer.writerow([
                po_common["vendor_name"],
                po_common["vendor_code"],
                po_common["po_number"],
                po_common["order_date"],
                po_common["delivery_date"],
                po_common["payment_term"],
                po_common["warehouse"],
                po_common["vendor_po_number"],
                po_common["order_date"],
                "",  # Vendor Invoice number
                "",  # Vendor Delivery Ref#
                po.invoice_date or "",
                "",  # Vendor Invoice Due Date
                "",  # Vendor Invoice Status
                "",  # Shipping Provider
                "",  # Tracking link
                item.product_id,              # Item SKU (replace if SKU exists)
                "",                           # Item Name (optional)
                item.qty,
                f"${item.price}",
                item.discount_percentage,
                item.tax_percentage,
                f"${item.tax_amount}",
                f"${item.subtotal}",
                f"${po_common['freight']}",
                f"${po_common['surcharge']}",
                po_common["comment"],
            ])

    return response

